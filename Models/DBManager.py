import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl import load_workbook

class DBManager:
    def __init__(self, nombre,cedula,celular,direccion,correo,producto,cantidad):
        self.nombre = nombre
        self.cedula = cedula
        self.celular = celular
        self.direccion = direccion
        self.correo = correo
        self.producto = producto
        self.cantidad = cantidad
        self.Datalist = {"Nombre": 0, "Cedula": 0, "Celular": 0, "Direccion":0, 
                         "Correo":0, "Producto":0, "Cantidad": 0, "Date":datetime.now().date()}
    
    def AlmacenarNombre (self):
        self.Datalist["Nombre"] = [self.nombre]
        self.Datalist["Cedula"] = [self.cedula]
        self.Datalist["Celular"] = [self.celular]
        self.Datalist["Direccion"] = [self.direccion]
        self.Datalist["Correo"] = [self.correo]
        self.Datalist["Producto"] = [self.producto]
        self.Datalist["Cantidad"] = [self.cantidad]
    
    
    def Write (self):
        self.Datalist = pd.DataFrame(self.Datalist)
        self.excel_file_path = "./Models/Database/DataBase.xlsx"
        self.DB = pd.read_excel(self.excel_file_path, sheet_name="Hoja1")
        self.DB = pd.concat([self.DB, self.Datalist], ignore_index=True)
        self.DB.to_excel(self.excel_file_path, sheet_name = "Hoja1",index=False)
        # Leer el archivo de Excel original
        df_origen = pd.read_excel(r'C:\Users\USUARIO\OneDrive - UCO\Erika y Mateo\Aplicativo\Models\DataBase\DataBase.xlsx')

        # Procesar los datos como desees
        df_nuevo = df_origen[['Producto', 'Cantidad', 'Date']]

        # Leer el archivo de Excel de destino
        archivo_destino = r'C:\Users\USUARIO\OneDrive - UCO\Erika y Mateo\Aplicativo\Models\DataBase\Existencias.xlsx'
        wb = load_workbook(filename=archivo_destino)
        ws = wb.active

        # Obtener la hoja de Excel destino (en este caso, la Hoja1)
        ws = wb['Hoja1']

        # Obtener las filas y columnas del DataFrame nuevo
        filas, columnas = df_nuevo.shape

        # Escribir los datos nuevos en el archivo de Excel de destino
        for i in range(filas):
            for j in range(columnas):
                ws.cell(row=i+2, column=j+1, value=df_nuevo.iloc[i, j])

        # Guardar el archivo de Excel de destino
        wb.save(archivo_destino)
        
        ##print("Write sucessfull")